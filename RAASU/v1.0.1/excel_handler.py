from openpyxl import Workbook, load_workbook
import os
from datetime import datetime

# Detect user's Documents folder
DOCUMENTS = os.path.join(os.path.expanduser("~"), "Documents")

SALES_DIR = os.path.join(DOCUMENTS, "raasu-pos-sales")


def get_today_file():

    today = datetime.now().strftime("%Y-%m-%d")

    if not os.path.exists(SALES_DIR):
        os.makedirs(SALES_DIR)

    return os.path.join(SALES_DIR, f"{today}.xlsx")


def save_invoice(data):

    file = get_today_file()

    if not os.path.exists(file):
        wb = Workbook()
        ws = wb.active
        ws.append(["Time", "Invoice", "Customer", "Qty", "Unit Price", "Total"])
        wb.save(file)

    wb = load_workbook(file)
    ws = wb.active

    for item in data["items"]:
        ws.append([
            datetime.now().strftime("%H:%M:%S"),
            data["invoice_no"],
            data["customer_name"],
            item["qty"],
            item["unit_price"],
            item["qty"] * item["unit_price"]
        ])

    wb.save(file)


def get_today_summary():

    file = get_today_file()

    if not os.path.exists(file):
        return 0, 0

    wb = load_workbook(file)
    ws = wb.active

    total_items = 0
    total_revenue = 0

    for row in ws.iter_rows(min_row=2, values_only=True):

        _, _, _, qty, _, total = row

        total_items += qty
        total_revenue += total

    return total_items, total_revenue